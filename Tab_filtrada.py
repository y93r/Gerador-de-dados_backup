import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Caminho do arquivo da planilha original e da matriz
arquivo_planilha = r'C:\Users\Usuário\Documentos\Cadastro_funcionarios\Informacoes.xlsx'
arquivo_matriz = r'C:\Users\Usuário\Documentos\Cadastro_funcionarios\Matriz.xlsx'

# Carrega a planilha original para um DataFrame
df_original = pd.read_excel(arquivo_planilha)

# Define a data limite no formato brasileiro (dd/mm/yyyy)
data_limite_str = '01/05/2023'
data_limite = datetime.strptime(data_limite_str, '%d/%m/%Y')

# Filtra os dados para manter apenas as linhas com datas menores que a data limite
dados_filtrados = df_original[pd.to_datetime(df_original['Data de Admissão'], format='%d/%m/%Y') < data_limite]

# Gera o nome do arquivo da matriz com base na data atual
data_atual = datetime.now().strftime('%Y-%m-%d')
arquivo_matriz_atual = (r'C:\Users\Usuário\Documents\MEGA\Backup_BD\Backup_{}.xlsx'.format(data_atual))

# Carrega a matriz existente ou cria uma nova planilha em branco com a mesma estrutura da planilha original
try:
    matriz = load_workbook(arquivo_matriz_atual)
except FileNotFoundError:
    matriz = load_workbook(arquivo_planilha)

# Seleciona a primeira aba da matriz
aba_matriz = matriz.active

# Limpa o conteúdo da matriz, mantendo apenas a formatação
aba_matriz.delete_rows(2, aba_matriz.max_row)

# Copia os dados filtrados para a matriz, preservando a formatação original
for row in dados_filtrados.itertuples(index=False):
    aba_matriz.append(row)

# Salva a matriz atualizada com o nome do arquivo correspondente à data atual
matriz.save(arquivo_matriz_atual)

print(f"Os dados filtrados foram salvos na matriz {arquivo_matriz_atual} com sucesso.")
